_author_ = 'Benjamin Krawitz, Chatan Konda'
_project_ = 'IRA Parser to Input into AppSec CSEQ Intake Form'

import openpyxl, pickle
import selenium.webdriver as webdriver
import selenium.webdriver.support.ui as ui
import pdb, os, json, sys, re, time
from simple_salesforce import Salesforce
from datetime import datetime
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def login(intakeNumber):
    driver = webdriver.Firefox(executable_path=os.getcwd() + '/geckodriver')
    driver.get('https://appsec.kp.org')
    time.sleep(3)
    driver.get('https://appsec.kp.org/auth')
    print "Please login to the website with your NUID and Windows Password! Then type 'continue'"
    return driver

def waitfor(reference, ByType):
	""" waits for an element using the By class"""
	delay = 2 # seconds
	try:
		myElem = WebDriverWait(driver, delay).until(EC.presence_of_element_located((ByType, reference)))
		# print "Element ready!"
		return myElem
	except TimeoutException:
		print "Loading timed out for" + reference

def cseq_update(CSEQ_MAP, driver):
    # Update REGULATORY SCOPE
    select = Select(driver.find_element_by_id('regulatoryScopeArr'))
    cseq_regulatory_arry = ['HIPAA', 'HIPAA Restricted', 'PCI', 'SOX', 'PII', 'None']
    for value in CSEQ_MAP['REGULATORY_SCOPE']:
        #making sure none is not included when other options are selected
        if value == 'None':
            break
        else:
            select.select_by_index(cseq_regulatory_arry.index('None'))
            select.select_by_index(cseq_regulatory_arry.index(value))

    time.sleep(2)

    #waitfor('//select[@id="numRecords"]//option[@value="' + CSEQ_MAP['MAX_NUM_RECORDS'] + '"]',By.XPATH).click()

    # Update NUM RECORDS
    driver.find_element_by_xpath('//select[@id="numRecords"]//option[@value="' + CSEQ_MAP['MAX_NUM_RECORDS'] + '"]').click()
    # Update Impact Care

    time.sleep(2)
    # Update Access Over Internet
    #driver.find_element_by_id('accessType').find_element_by_xpath('//option[@value="' + CSEQ_MAP['INTERNET_USE'] + '"]').click()

    time.sleep(2)
    # Update Class Of Service
    driver.find_element_by_id('clzOfService').find_element_by_xpath('//option[@value="' + CSEQ_MAP['COS'].split(":")[0] + '"]').click()

    time.sleep(2)
    # Update KP Branded
    driver.find_element_by_id('kpBrandedYn').find_element_by_xpath('//option[@value="' + CSEQ_MAP['BRANDED'] + '"]').click()
    time.sleep(2)
    # Update Authentication
    driver.find_element_by_id('authenticatedYn').find_element_by_xpath('//option[@value="' + CSEQ_MAP['AUTH'] + '"]').click()
    time.sleep(2)
    # Update type of technology
    select = Select(driver.find_element_by_id('appTypeArr'))
    cseq_system_tech = ['Web Application', 'Web Service', 'Mobile Application', 'Other with available source code', 'Other with no source code available', 'Not an Application']
    # select.select_by_index(cseq_system_tech.index('Web Service'))
    if len(CSEQ_MAP['SYSTEM_TECH']) == 0:
        print("Array is Empty")
    else:
        #select.select_by_index(cseq_system_tech.index('Not an Application'))
        for value in CSEQ_MAP['SYSTEM_TECH']:
            select.select_by_index(cseq_system_tech.index(value))

    time.sleep(2)
    # Update Pilot or POC
    driver.find_element_by_id('pilotOrPocEval').find_element_by_xpath('//option[@value="' + CSEQ_MAP['PILOT'] + '"]').click()

    time.sleep(2)
    # Update App Development Origin
    driver.find_element_by_id('appDevelopmentOrigin').find_element_by_xpath('//option[@value="' + CSEQ_MAP['DEVELOPED'] + '"]').click()

    time.sleep(2)
    # Update data center
    select = Select(driver.find_element_by_id('dataCenterArr'))
    cseq_data_center = ['Boulder', 'Corona', 'Downey', 'Irvine', 'KP Private Cloud (Softlayer/Bluemix)', 'Napa', 'Pleasanton', 'Silver Springs', 'Walnut Creek', 'Not Hosted at Kaiser Permanente']

    for value in CSEQ_MAP['DATA_CENTER']:
        select.select_by_index(cseq_data_center.index(value))
    #select.select_by_index(cseq_data_center.index('Not Hosted at Kaiser Permanente'))

    time.sleep(2)
    # Update SDLC Phase
    driver.find_element_by_id('sdlcPhase').find_element_by_xpath('//option[@value="' + CSEQ_MAP['SDLC'] + '"]').click()

    time.sleep(2)
    # Update Application Location
    driver.find_element_by_id('applicationLocation').find_element_by_xpath('//option[@value="' + CSEQ_MAP['APPLICATION_LOCATION'] + '"]').click()

    time.sleep(2)
    # Update ASG Support
    driver.find_element_by_id('asgYn').find_element_by_xpath('//option[@value="' + CSEQ_MAP['ASG_SUPPORT'] + '"]').click()

    driver.find_element_by_id('impactCareYn').find_element_by_xpath('//option[@value="' + CSEQ_MAP['IMPACT_PATIENT_CARE'] + '"]').click()

    driver.find_element_by_id('accessType').find_element_by_xpath('//option[@value="' + CSEQ_MAP['INTERNET_USE'] + '"]').click()


    # Update Go Live Date
    #Error here? cant clear
    pdb.set_trace()
    goLive = driver.find_element_by_id('goLiveDate')
    goLive.clear()
    goLive.send_keys(CSEQ_MAP['GO_LIVE_DATE'])

    print "******** DONE PROCESSING : REVISE!!! *********"
    pdb.set_trace()

    # HIT SUBMIT


def parseResults(intakeNumber, driver):
    driver.switch_to_window(driver.window_handles[-1])
    time.sleep(1)
    team_box = driver.find_elements_by_xpath('//ul[@class="category-list slide"]')[:-1]
    payload = {
        "Cyber_Security_Business_Impact__c" : driver.find_element_by_xpath('//div[@class="post-content"]//b').text,
        "Initial_AppSec_Applicability__c" : "Applicable" if "Assessment required. (" in team_box[0].text else "Not Applicable",
        "Initial_Security_Arch_Applicability__c" : "Applicable" if "Assessment required." in team_box[1].text else "Not Applicable",
        "Initial_Red_Team_Applicability__c" : "Applicable" if "Assessment required." in team_box[2].text else "Not Applicable"
    }

    appsec_priority = [ele for ele in driver.find_elements_by_xpath('//ul[@class="category-list slide"]')[:-1][0].text if ele.isdigit()]
    if len(appsec_priority) == 1:
        payload['Initial_AppSec_Priority__c'] = "P" + str(appsec_priority[0])

    #sf = Salesforce(username="benjamin.krawitz@kp.org.prod.devqa", password="", security_token="VQ5zciyh35rBZV723nW2EJj2", sandbox=True)
    sf = "Salesforce()"
    res = sf.query("SELECT Id, TRO_Risk_Engine_Intake_Number__c from Inherent_Risk_Assessment__c where TRO_Risk_Engine_Intake_Number__c = '%s'" % intakeNumber)
    id_to_update = res['records'][0]['Id']
    pdb.set_trace()
    response = sf.Inherent_Risk_Assessment__c.update(id_to_update,payload)
    if response.code == 204:
        print "Success updating " + id_to_update
    else:
        print "Error updating " + id_to_update

def main():
    i = 0
    for filename in os.listdir(os.getcwd() + "/PROCESS_CSEQ"):
        if filename.split(".")[-1] == "xlsm":
            intakeNumber = filename.split(".")[0]
            print "Parsing " + intakeNumber + " file starting..."
            if i ==0:
                driver = login(intakeNumber)
                time.sleep(3)
            pdb.set_trace()
            driver.find_element_by_id('intakeNumberForm').find_element_by_tag_name('input').send_keys(filename.split(".")[0])

            time.sleep(3)
            driver.find_element_by_xpath('//input[@value="Go!"]').click()
            CSEQ_MAP = parseIRA(intakeNumber)
            pdb.set_trace()
            print CSEQ_MAP
            cseq_update(CSEQ_MAP, driver)
            parseResults(intakeNumber, driver)
            i+=1

def parseIRA(intakeNumber):
    #Step 1: Open the xlsm workbook and pull the questionaire sheet which is the by default the 1st sheet in the workbook
    wb = openpyxl.load_workbook(os.getcwd() + "/PROCESS_CSEQ/" + intakeNumber+".xlsm")
    #Step 2: By default active method pulls first worksheet
    ws = wb.active
    # Loop through each file and make a JSON object of all CSEQ questions
    CSEQ_MAP = dict(
        TRO_INTAKE_NUM = intakeNumber,
        REGULATORY_SCOPE = regulatory_Scope(ws),
        MAX_NUM_RECORDS = max_Num_Records(ws),
        IMPACT_PATIENT_CARE = impact_Patient_Care(ws),
        INTERNET_USE = internet_Use(ws),
        COS = class_Of_Service(ws),
        BRANDED = kp_Branded(ws),
        AUTH = auth_Req(ws),
        SYSTEM_TECH = system_Tech(ws),
        PILOT = poc_Pilot(ws),
        DEVELOPED = dev_Team(ws),
        DATA_CENTER = data_Centers(ws),
        SDLC = sdlc_Phase(ws),
        APPLICATION_LOCATION = app_Loc(ws),
        ASG_SUPPORT = asg_Support(ws),
        GO_LIVE_DATE = go_Live_Date(ws)
    )
    return CSEQ_MAP

# Pull REGULATORY_SCOPE from IRA
def regulatory_Scope(ws):
    regulatoryScope = []

    #HIPAA
    if ws['A195'].value == True:
        regulatoryScope.append("HIPAA")
    #Restricted HIPAA
    if ws['A168'].value == True:
        regulatoryScope.append("HIPAA Restricted")
    #PCI
    if ws['A136'].value == True or ws['A205'].value == True:
        regulatoryScope.append('PCI')
    #SOX
    if ws['A161'].value == True or ws['A204'].value == True:
        regulatoryScope.append('SOX')
    #PII
    if ws['A123'].value == True:
        regulatoryScope.append('PII')
    #None
    #if regulatory array is empty
    if not regulatoryScope:
        regulatoryScope.append('None')

    print regulatoryScope
    return regulatoryScope

# Pull MAX_NUM_RECORDS affected by app/service from IRA
def max_Num_Records(ws):
    # Create an array of cells to loop through for getting max recs affected
    maxRecsBooleanCells = []

    # for loops range (inclusive, exclusive)
    #Internal PHI
    for i in range(94, 101):
        maxRecsBooleanCells.append('A' + str(i))

    #External PHI
    for i in range(102, 109):
        maxRecsBooleanCells.append('A' + str(i))

    #PII
    for i in range(126, 132):
        maxRecsBooleanCells.append('A' + str(i))

    #PCI
    for i in range(139, 145):
        maxRecsBooleanCells.append('A' + str(i))

    maxRecs = 0
    recs = "None"
    for i in maxRecsBooleanCells:
        if ws[i].value == True:
            if maxRecs < 499 and (i == 'A95' or i == 'A103'or i == 'A127' or i == 'A140'):
                maxRecs = 499
                recs = "Department 100's"
            elif maxRecs < 50000 and (i == 'A95' or i == 'A103'or i == 'A127' or i == 'A140'):
                maxRecs = 50000
                recs = "Facility or Service area wide 1K's"
            elif maxRecs < 100000 and (i == 'A97' or i == 'A105'):
                maxRecs = 100000
                recs = "Facility or Service area wide 1K's"
            elif maxRecs < 500000 and (i == 'A98' or i == 'A106'):
                maxRecs = 500000
                recs = "Region-wide 100K's"
            elif maxRecs < 10000000 and (i == 'A99' or i == 'A107'):
                maxRecs = 10000000
                recs = "National 1M's"
            elif maxRecs < 10000001 and (i == 'A100' or i == 'A108'):
                maxRecs = 10000001
                recs = "National 1M's"

    print "Maximum Records Affected = " + recs
    return recs


# Pull from IRA App/Service IMPACTS_PATIENT_CARE Boolean
def impact_Patient_Care(ws):
    print "Impact Patient Care? " + str(ws['A253'].value)
    return "No" #if ws['A253'].value == True else "No" #changed to test


# Pull from IRA App/Service INTERNET_USE Boolean
def internet_Use(ws):
    print "Internet Usable? " + str(ws['A303'].value)
    return "Yes" if ws['A303'].value == "True" else "No"

# Pull Class Of Service from IRA
def class_Of_Service(ws):
    cos_cnt = 8 #OG: cos_cnt = 10
    cnt = 0
    for i in range(226, 234):
        cell = 'A' + str(i)
        if ws[cell].value == True and cos_cnt == 8: #OG: cos_cnt = 10
            cos_cnt = cnt
        cnt += 1

    cos = str(cos_Converter(cos_cnt))

    print "CLASS OF SERVICE = " + cos
    return cos

#Convert COS int val from IRA to appropriate string on AppSec CSEQ Intake Form
def cos_Converter(cos):
    # Essentially a switch statement for converting COS int val into appropriate appsec string answer
    return {
        0: "0: Recovery Time Objective is 0 hours (High Availabilty)",
        1: "1: Recovery Time Objective is <= 4 hours",
        2: "2: Recovery Time Objective is <= 24 hours",
        3: "3: Recovery Time Objective is <= 72 hours",
        4: "4: Recovery Time Objective is <= 1 week",
        5: "5: Recovery Time Objective is <= 2 weeks",
        6: "6: Recovery Time Objective is <= 1 month",
        7: "7: Recovery Time Objective is best effort",
        8: "8: No Recovery"
    }[cos]

# Pulls KP_BRANDED Boolean from IRA
def kp_Branded(ws):
    print "Branded? " + str(ws['A417'].value)
    return "Yes" if ws['A417'].value == True else "No"

# Pulls AUTH Boolean from IRA
def auth_Req(ws):
    print "Authentication Required? " + str(ws['A306'].value)
    return "Yes" if ws['A306'].value == True else "No"


# Pulls list of SYSTEM_TECH associated with App/Service from IRA
def system_Tech(ws):
    system_Tech = {}

    for i in range(62, 65):
        bool_cell = 'A' + str(i)
        val_cell = 'H' + str(i)
        bool = ws[bool_cell].value
        val = ws[val_cell].value
        if bool == True:
            system_Tech[str(val)] = True
        else:
            system_Tech[str(val)] = False

    if ws['A66'].value == True:
        system_Tech["Other with available source code"] = True
    else:
        system_Tech["Other with available source code"] = False

    if ws['A67'].value == True:
        system_Tech["Other with no source code available"] = True
    else:
        system_Tech["Other with no source code available"] = False

    sysTech = []

    for i in system_Tech:
        if system_Tech[i] == True:
            sysTech.append(i)
        # if len(sysTech) == 0:
        #     sysTech.append("Not an Application")

    print "System Tech: "
    print sysTech
    return sysTech


# Pull PILOT/POC Phase from IRA
def poc_Pilot(ws):
    proj_type = ws['H29'].value

    if proj_type.upper() == "PILOT":
        pilot = "Pilot"
    elif proj_type.upper() == "POC / EVALUATION":
        pilot = "POC/Evaluation"
    else:
        pilot = "None"

    print "PROJECT PHASE = " + pilot
    return pilot

# Pulls what type of team DEVELOPED app/service from IRA
def dev_Team(ws):
    #default developers value in case of error
    developers = "Unknown"

    # for loops range (inclusive, exclusive)
    for i in range(71, 75):
        bool = 'A'+ str(i)
        val = 'H' + str(i)
        if ws[bool].value == True:
            developers = dev_Team_Converter(i)

    #Error Check for Development Team
    if developers == "Unknown":
        print "Error: Please review cells K71 - K74 in the excel file!!!"

    print "Development Team: " + developers
    return developers

def dev_Team_Converter(i):
    return {
        71: "Developed Internally by KP",
        72: "Developed by external outsourced team",
        73: "Vendor Product customized for KP use",
        74: "Vendor Product for general public use"
    }[i]



# Pull list of DATA_CENTER(s) App/Service is hosted on from IRA
def data_Centers(ws):

    dcList = []
    cnt = 0
    for i in range (238, 245):
        bool = 'A'+str(i)
        if ws[bool].value == True:
            dcList.append(dc_Converter(i))
            cnt +=1

    if cnt == 0:
        dcList.append("Not Hosted at Kaiser Permanente")

    # No Option for DOWNEY OR SOFTLAYER/BMX
    print "Data Centers Hosting Application Solution: " + str(dcList)
    return dcList



def dc_Converter(i):
    return {
        238: "Boulder",
        239: "Corona",
        240: "Irvine",
        241: "Napa",
        242: "Pleasanton",
        243: "Silver Springs",
        244: "Walnut Creek"
    }[i]

# Pull SDLC phase from IRA
def sdlc_Phase(ws):
    sdlc_stage = ws['H31'].value

    sdlc_stage = 'Testing' if 'Testing' in sdlc_stage else sdlc_stage

    if 'N/A' in sdlc_stage:
        sdlc_stage = 'N/A'

    print "SDLC Stage = " + sdlc_stage
    return sdlc_stage

# Pull APPLICATION_LOCATION from IRA
def app_Loc(ws):
    appLoc = 'Unknown(E.g. In a location with unknown security controls)'
    appLocList = []
    if ws['A236'].value == True:
        appLoc = '3rd Party Hosted (Overseas)'
        appLocList.append('3rd Party Hosted (Overseas)')
    if ws['A237'].value == True:
        appLoc = '3rd Party Hosted (Domestic)'
        appLocList.append('3rd Party Hosted (Domestic)')

    kpit_facility = False

    for i in range (238, 250):
        bool = 'A' + str(i)
        if ws[bool].value == True:
            kpit_facility = True

    if kpit_facility == True:
        if ws['A76'].value == True:
            appLoc = 'Inside KP facilities managed by KP IT'
            appLocList.append('Inside KP facilities managed by KP IT')
        else:
            appLoc = 'Inside KP facilities but not managed by KP IT'
            appLocList.append('Inside KP facilities but not managed by KP IT')

    if len(appLocList) == 0:
        appLocList.append(appLoc)

    print "Application Location: " + appLoc
    print appLocList
    if len(appLocList) > 1:
        pdb.set_trace()
    else:
        return appLocList[0]


# Pulls if App/Service has ASG_SUPPORT from IRA
def asg_Support(ws):
    support = False
    if ws['A78'].value == True:
        print("ASG Support: True")
    else:
        print("ASG Support: False")
    return "Yes" if ws['A78'].value == True else "No"

# Pulls GO_LIVE_DATE from IRA
def go_Live_Date(ws):
    #goLive = ws['H16'].value.strftime('%Y-%m-%d')
    liveDate = str(ws['H16'].value)
    #print(liveDate)
    goLiveTest = None



    if liveDate.find(".") != -1:
        goLiveTest = datetime.strptime(liveDate, "%Y-%m-%d %H:%M:%S")
    elif liveDate.find("/") != -1:
        goLiveTest = datetime.strptime(liveDate, "%Y-%m-%d %H:%M:%S")
    elif liveDate.find("-") != -1:
        goLiveTest = datetime.strptime(liveDate, "%Y-%m-%d %H:%M:%S")
    elif liveDate.find("_") != -1:
        goLiveTest = datetime.strptime(liveDate, "%Y-%m-%d %H:%M:%S")
    elif liveDate.find(",") != -1:
        goLiveTest = datetime.strptime(liveDate, "%Y-%m-%d %H:%M:%S")

    print(goLiveTest) #goLive

    return goLiveTest #newDate #goLive

main()
